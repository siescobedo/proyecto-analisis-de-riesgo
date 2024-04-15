import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl

class AnalisisDeRiesgoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Analisis de Riesgo")
        self.root.geometry('350x300')

        self.root.configure(bg='#ECECEC')
        style = ttk.Style()
        style.theme_create('custom', settings={
            'TLabel': {'configure': {'background': '#EC0000', 'foreground': '#FFFFFF', 'font': ('Helvetica', 11)}},
            'TFrame': {'configure': {'background': '#EC0000', 'borderwidth': 0, 'relief': 'flat'}},
            'TButton': {'configure': {'background':'#DEEDF2', 'font': ('Helvetica', 11), 'anchor': 'center'}}}
            )

        style.theme_use('custom')

        frame = ttk.Frame(root)
        frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(frame, text="Analisis de Riesgo", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10, sticky='en')

        ttk.Label(frame, text="Archivo internos:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_internos, cursor="hand2").grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(frame, text="Archivo externos:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_externos, cursor="hand2").grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(frame, text="Archivo RACF:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_racf, cursor="hand2").grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(frame, text="Archivo Analisis:").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_analisis, cursor="hand2").grid(row=4, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(frame, text="Ejecutar", command=self.procesar,cursor="hand2", width=20).grid(row=6, column=1, columnspan=2, pady=10, sticky='n')

        self.internos_procesado = ""
        self.externo_procesado = ""
        self.racf_procesado = ""
        self.analisis_procesado = ""
    
    def cargar_archivo(self, tipo):
        file_path = filedialog.askopenfilename(title=f"Seleccionar archivo {tipo.capitalize()}", filetypes=[(f"{tipo} files", f"*.{tipo}")])
        return file_path

    def cargar_internos(self):
        self.internos_path = self.cargar_archivo("xlsx")

    def cargar_externos(self):
        self.externos_path = self.cargar_archivo("xlsx")
    
    def cargar_racf(self):
        self.racf_path = self.cargar_archivo("xlsx")

    def cargar_analisis(self):
        self.analisis_path = self.cargar_archivo("xlsx")
    
    def procesar(self):
        if self.internos_path is not None and self.externos_path and self.racf_path is not None and self.analisis_path:
            self.procesar_internos()
            self.procesar_externos()
            self.procesar_racf()
            self.hacer_analisis()
        else:
            tk.messagebox.showwarning("Error", "Debe seleccionar todos los archivos.")

    def procesar_internos(self):
        wb_i = openpyxl.load_workbook(self.internos_path)
        ws_i = wb_i.active

        internos_c = {}
        for cell in ws_i[1]:
            if cell.value != None:
                internos_c[cell.value] = cell.column_letter
        
        self.internos_dict = {}
        for i in range(2, ws_i.max_row+1):
            # get data
            rut = ws_i[f"{internos_c['Rut']}{i}"].value
            cargo = ws_i[f"{internos_c['Cargo']}{i}"].value
            gls_cargo = ws_i[f"{internos_c['GlsCargo']}{i}"].value
            unirel = ws_i[f"{internos_c['UR']}{i}"].value
            gls_unirel = ws_i[f"{internos_c['GlsUR']}{i}"].value
            nombre_jefe = ws_i[f"{internos_c['NombreJefe']}{i}"].value

            # make new intern dic'
            interno = {"Cargo": cargo, "GlsCargo": gls_cargo,
                    "Unirel": unirel, "GlsUnirel": gls_unirel,
                    "NombreJefe": nombre_jefe}
            
            # add to interns dict
            self.internos_dict[rut] = interno

    def procesar_externos(self):
        wb_e = openpyxl.load_workbook(self.externos_path)
        ws_e = wb_e.active

        externos_c = {}
        for cell in ws_e[1]:
            if cell.value != None:
                externos_c[cell.value] = cell.column_letter
        
        self.externos_dict = {}
        for i in range(2, ws_e.max_row+1):
            # get data
            rut = ws_e[f"{externos_c['Rut']}{i}"].value
            supervisor = ws_e[f"{externos_c['Supervisor Externo']}{i}"].value

            # make new external dict
            external = {"Supervisor Externo": supervisor}
            
            # add to externos dict
            self.externos_dict[rut] = external

    def procesar_racf(self):
        wb_r = openpyxl.load_workbook(self.racf_path)
        ws_r = wb_r.active

        racf_c = {}
        for cell in ws_r[1]:
            if cell.value != None:
                racf_c[cell.value] = cell.column_letter
        
        self.racf_dict = {}
        for i in range(2, ws_r.max_row+1):
            # get data
            rut = ws_r[f"{racf_c['Rut']}{i}"].value
            account_name = ws_r[f"{racf_c['USBD_NAME']}{i}"].value

            # add to racf dict
            self.racf_dict[account_name] = rut

    def hacer_analisis(self):
        wb_a = openpyxl.load_workbook(self.analisis_path)
        ws_a = wb_a.active

        last_column_index = ws_a.max_column
        new_column_names = ["Rut", "Cargo", "GlsCargo", "UR", "GlsUR", "NombreJefe", "Supervisor Externo"]
        for index, name in enumerate(new_column_names, start=1):
            ws_a.cell(row=1, column=last_column_index + index, value=name)


        analisis_c = {}
        for cell in ws_a[1]:
            if cell.value != None:
                analisis_c[cell.value] = cell.column_letter

        for i in range(2, ws_a.max_row+1):
            account_name = ws_a[f"{analisis_c['Nombre de la cuenta']}{i}"].value
            if account_name in self.racf_dict.keys():
                
                rut = self.racf_dict[account_name]
                if rut in self.internos_dict.keys():
                    data = self.internos_dict[rut]
                    ws_a[f"{analisis_c['Rut']}{i}"] = rut
                    ws_a[f"{analisis_c['Cargo']}{i}"] = data["Cargo"]
                    ws_a[f"{analisis_c['GlsCargo']}{i}"] = data["GlsCargo"]
                    ws_a[f"{analisis_c['UR']}{i}"] = data["UR"]
                    ws_a[f"{analisis_c['GlsUR']}{i}"] = data["GlsUR"]
                    ws_a[f"{analisis_c['NombreJefe']}{i}"] = data["NombreJefe"]
                elif rut in self.externos_dict.keys():
                    data = self.externos_dict[rut]
                    ws_a[f"{analisis_c['Rut']}{i}"] = rut
                    ws_a[f"{analisis_c['Supervisor Externo']}{i}"] = data["Supervisor Externo"]
                else:
                    print(f"Rut {rut} not found in Internos or Externos")
            else:
                print(f"Account Name {account_name} not found in RACF")
        
        ws_a.save(self.analisis_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = AnalisisDeRiesgoApp(root)
    root.mainloop()